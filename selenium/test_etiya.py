from selenium import webdriver
from selenium.webdriver.common.by import By
import pytest
import openpyxl
# debugging
# breakpoint => kodu debug modda iken bekleteceğimiz satır

# @pytest.mark.skip


class TestEtiya():
    # setup_method => pytest tarafından tanınan
    # bulunduğunda her test öncesi otomatik çalıştırılan fonk.
    def setup_method(self, method):
        # her test başlangıcında çalışacak fonksiyon
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get("https://www.saucedemo.com/")
    # teardown_method

    def teardown_method(self, method):
        # her test bitiminde çalışacak fonksiyon
        self.driver.quit()

    def readDataFromExcel():
        excelFile = openpyxl.load_workbook("userFailData.xlsx")
        sheet = excelFile["Sheet1"]
        rows = sheet.max_row  # sayfadaki max satır sayısını alır.
        data = []
        for i in range(1, rows):
            username = sheet.cell(i, 1).value
            password = sheet.cell(i, 2).value
            data.append((username, password))
        return data

    def test_header(self):
        logo = self.driver.find_element(By.CLASS_NAME, 'login_logo')
        assert logo.text == "Swag Labs"

    # bu test aynı anda 3 veriyle çalışsın.
    # abc-123  deneme-12345  etiya-etiya
    # annotation
    @pytest.mark.parametrize("username,password", readDataFromExcel)
    def test_login_invalid(self, username, password):
        loginBtn = self.driver.find_element(By.ID, 'login-button')
        usernameInput = self.driver.find_element(By.ID, "user-name")
        passwordInput = self.driver.find_element(By.ID, "password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn.click()
        errorContainer = self.driver.find_element(
            By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorContainer.text == "Epic sadface: Username and password do not match any user in this service"
# prefix => ön ek
